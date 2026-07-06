"""Shared XLSX reading utilities."""
from __future__ import annotations

import io
from pathlib import Path
from typing import BinaryIO


def read_xlsx_sheets(
    source: Path | BinaryIO,
) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    if isinstance(source, Path):
        wb = load_workbook(source, read_only=True, data_only=True)
    else:
        wb = load_workbook(source, read_only=True, data_only=True)

    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)
            if rows:
                sheets.append((sheet_name, rows))
    finally:
        wb.close()
    return sheets


def sheet_to_table_text(sheet_name: str, rows: list[list[str]]) -> str:
    lines = [f"# {sheet_name}", "[TABLE]"]
    for row in rows:
        lines.append(" | ".join(row))
    lines.append("[/TABLE]")
    return "\n".join(lines)
